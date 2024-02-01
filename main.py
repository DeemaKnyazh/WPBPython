import base64
import qrcode
from qrcode.image.styledpil import StyledPilImage
import os
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName, FileType, Disposition)
from qrcode.image.styles.colormasks import SolidFillColorMask
from PIL import Image, ImageDraw
import openpyxl
import math
import requests
import pdfkit
from xhtml2pdf import pisa  # import python module
from dotenv import load_dotenv
import os

load_dotenv()
wb = openpyxl.load_workbook('GuestListTest.xlsx')
sheet = wb.active
api_url = "https://ruskokaaccess.azurewebsites.net/api/guests"
headers = {"Authorization": os.environ.get("api-key")}
loca = os.environ.get('loc')

# ToDo
# Improve PDF
# Remove generic Table number when adding to DB
def convert_html_to_pdf(html_content, pdf_path):
    try:
        pdfkit.from_string(html_content, pdf_path)
        print(f"PDF generated and saved at {pdf_path}")
    except Exception as e:
        print(f"PDF generation failed: {e}")


path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)


def get_image_file_as_base64_data(FILEPATH):
    with open(FILEPATH, 'rb') as image_file:
        return base64.b64encode(image_file.read())


def style_eyes(img):
    img_size = img.size[0]
    eye_size = 70  # default
    quiet_zone = 40  # default
    mask = Image.new('L', img.size, 0)
    draw = ImageDraw.Draw(mask)
    draw.rectangle((10, 10, 80, 80), fill=255)
    draw.rectangle((img_size - 80, 10, img_size - 10, 80), fill=255)
    draw.rectangle((10, img_size - 80, 80, img_size - 10), fill=255)
    return mask


for i in range(sheet.max_row - 2):
    emailGroup = [[0, 0, 0, 0, 0]]
    guests = 1
    names = [0]
    namesUse = [0]
    first = sheet.cell(row=3 + i, column=1)
    last = sheet.cell(row=3 + i, column=2)
    emails = sheet.cell(row=3 + i, column=4)
    status = sheet.cell(row=3 + i, column=5)

    names[0] = first.value + last.value
    namesUse[0] = first.value + " " + last.value

    emailGroup[0][0] = first.value
    emailGroup[0][1] = last.value
    emailGroup[0][2] = emails.value
    emailGroup[0][3] = status.value

    if emailGroup[0][3] == False:
        sheet.cell(row=3 + i, column=5).value = True
        for l in range((sheet.max_row - 3) - i):
            firstCheck = sheet.cell(row=4 + i + l, column=1)
            lastCheck = sheet.cell(row=4 + i + l, column=2)
            emailsCheck = sheet.cell(row=4 + i + l, column=4)
            statusCheck = sheet.cell(row=4 + i + l, column=5)
            if emailsCheck.value == emailGroup[0][2] and statusCheck.value == False:
                emailGroup.append([0, 0, 0, 0, 0])
                sheet.cell(row=4 + i + l, column=5).value = True
                names.append(firstCheck.value + lastCheck.value)
                namesUse.append(firstCheck.value + " " + lastCheck.value)
                emailGroup[l + 1][0] = firstCheck.value
                emailGroup[l + 1][1] = lastCheck.value
                emailGroup[l + 1][2] = emailsCheck.value
                emailGroup[l + 1][3] = statusCheck.value

        print(first.value, last.value, emails.value, emailGroup[0][3])
        print(emailGroup)

        for index, name in enumerate(names):
            # Generates the ticket code
            a = emailGroup[index][0][0]
            b = emailGroup[index][1][0]
            c = str((math.ceil((len(emailGroup[index][1] + emailGroup[index][0])) / 3.141592) * 2))
            d = emailGroup[index][0][-1]
            e = emailGroup[index][1][-1]
            code = ("WPB" + a + b + c + d + e + "23")
            emailGroup[index][4] = code

            # Generate a qr code based off the ticket code
            qr = qrcode.QRCode(version=5, error_correction=qrcode.constants.ERROR_CORRECT_H, border=1)
            qr.add_data(code)
            qr_eyes_img = qr.make_image(image_factory=StyledPilImage,
                                        color_mask=SolidFillColorMask(back_color=(255, 255, 255),
                                                                      front_color=(158, 42, 43)))
            qr_img = qr.make_image(image_factory=StyledPilImage,
                                   color_mask=SolidFillColorMask(back_color=(207, 234, 250), front_color=(84, 11, 14)))

            mask = style_eyes(qr_img)
            final_img = Image.composite(qr_eyes_img, qr_img, mask)
            final_img.save(emailGroup[index][0] + emailGroup[index][1] + ".png")

        # Creating a pdf
        trs = []
        for index, name in enumerate(namesUse):
            trs.append(f'''\
                          <tr>
                            <td align="center" size="bigger">{name}</td>
                          </tr>
                          <tr>
                            <td align="center"><img src="{loca}{names[index]}.png" style="zoom:90%" align="middle"></td>
                          </tr>
                          <tr>
                            <td><br></td>
                          </tr>
                          <tr>
                            <td><br></td>
                          </tr>
                          <tr>
                            <td><br></td>
                          </tr>
                          <tr>
                            <td style="margin: 30px"><br></td>
                          </tr>
                          ''')
        name_pdf_table = '\n'.join(trs)

        html = f"""\
        <html>
            <body style="font-size:20px;background-color:#CFEAFA">
                <div align="center">
                         <img width="90%" src='{loca}WPB-removebg.png'>
                         <p>Thank you for purchasing a ticket and supporting Ruskoka Camp!<br></p>
                         <table border="0" cellspacing="0" cellpadding="0">
                         <tr>
                            <td><br></td>
                          </tr>
                          <tr>
                            <td><br></td>
                          </tr>
                             {name_pdf_table}
                         </table>
                </div>
            </body>
        </html>
        """.format(name_pdf_table=name_pdf_table, loca=os.environ.get('loc'))
        print(html)
        pdfkit.from_string(html, names[0] + '.pdf', configuration=config, options={"enable-local-file-access": ""})

        # Send Email
        sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
        message = Mail(from_email='deema@ruskoka.com',
                       to_emails=emails.value)
        namess = [0]
        namess[0] = emailGroup[0][0] + emailGroup[0][1]
        print(namess)
        for index, name in enumerate(namess):
            with open(name + ".pdf", 'rb') as pdf:
                data = pdf.read()
                pdf.close()
            encoded_file = base64.b64encode(data).decode()

        attachedFile = Attachment(
            FileContent(encoded_file),
            FileName('tickets.pdf'),
            FileType('application/pdf'),
            Disposition('attachment')
        )
        message.attachment = attachedFile
        message.dynamic_template_data = {}
        message.template_id = os.environ.get('TEMPLATE')
        response = sg.send(message)
        print(response.status_code, response.body, response.headers)

        print("Message sent!")

        for index, name in enumerate(names):
            guest = {"name": emailGroup[index][0] + " " + emailGroup[index][1], "tables": 1,
                     "ticket": emailGroup[index][4]}
            response = requests.post(api_url, json=guest, headers=headers)
            print(response)

# Saves the workbook after each success is set to true
# wb.save('GuestListTest.xlsx')

# smtp.quit()  # finally, don't forget to close the connection
